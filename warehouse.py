import pandas as pd
from database.static import *
import numpy as np
from warehouse import DateLabels
from openpyxl import Workbook
import openpyxl
from utils import *


def loadTurboExcelFiles(fromYear, toYear):
    for y in range(fromYear, toYear+1):
        df = pd.read_csv("database/clean-data/DSE-"+str(y)+".csv", low_memory=False)
        for c in all:
            print("Year "+ str(y) +" Scrip "+c)
            filtered_df = df[df['Scrip'] == c]
            if filtered_df.empty:
                continue
            filtered_df = filtered_df.sort_values(by=['DayIndex'])
            applyMultipleTurbo(company=c, year=y, dataframe=filtered_df, turboX=[1,3,5,7,11,22,44])


def applyMultipleTurbo(company, year, dataframe, turboX):
    print(company)
    for turbo in turboX:
        print("Turbo value "+str(turbo))
        weights = np.array([0.0]*366)
        turbo -= 1
        maxIndex = dataframe.index.max()
        maxIndex = int(maxIndex)
        for startIndex in dataframe.index:
            endIndex = startIndex + turbo
            endIndex = int(endIndex)
            #print(maxIndex)
            if endIndex > maxIndex:
                break
            dayIndex1 = int(dataframe.loc[startIndex, 'DayIndex'])
            dayIndex2 = int(dataframe.loc[endIndex, 'DayIndex'])
            close1 = float(dataframe.loc[startIndex, 'Close'])
            close2 = float(dataframe.loc[endIndex, 'Close'])
            open1 = float(dataframe.loc[startIndex, 'Open'])
            open2 = float(dataframe.loc[endIndex, 'Open'])
            high1 = float(dataframe.loc[startIndex, 'High'])
            high2 = float(dataframe.loc[endIndex, 'High'])
            low1 = float(dataframe.loc[startIndex, 'Low'])
            low2 = float(dataframe.loc[endIndex, 'Low'])
            high2 = float(dataframe.loc[endIndex, 'High'])
            rr = 0
            #calculateTotalVolume()
            if turbo == 0:
                turbo = 1
            if close2 > open1 and close2 > high1:
                # apply bullish logic 1.3x
                rr = (close2-open1)/open1*130.0/turbo
            elif close2 < open1 and close2 < low1:
                # apply bearish 1.3x
                rr = (close2-open1)/open1*130.0/turbo
            elif close2 > open1:
                rr = (close2-open1)/open1*80.0/turbo # bullish 0.7x
            elif close2 < open1:
                rr = (close2-open1)/open1*80.0/turbo #bearish 0.7x
            #print(rr)
            for d in range(dayIndex1, dayIndex2+1):
                #print(weights[d]+ rr)
                weights[d] = round(rr, 4)+weights[d]
        #dbCon = connectDB()
        #print(weights, rr)
        startL = company[0]
        path = "C:/Users/User/Documents/warehouse/"+str(turbo)+"/"+startL+".xlsx"
        ws = None
        df = None
        wb = None
        
        #pandas.ExcelWriter(path, date_format=None, mode=’w’)
        try:
            df = pd.read_excel(path, sheet_name=company)
        except:
            df = pd.DataFrame({"Date": DateLabels})
        #print(df)
        #print(wb.sheetnames)
        #print(df)
        # no columns yet, set 1st Date col
        
        with pd.ExcelWriter(path, mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
            df[str(year)] = weights[1:366]
            # use to_excel function and specify the sheet_name and index to
            df.to_excel(writer, index=False, sheet_name=company)



loadTurboExcelFiles(2010,2022)
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# ws.title = "My sheet name"
# #ws2 = wb.create_sheet("Another Name", 0)
# wb.save("Test.xlsx")


